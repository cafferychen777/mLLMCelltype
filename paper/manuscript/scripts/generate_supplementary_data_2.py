#!/usr/bin/env python3
"""
Generate Supplementary Data 2 Excel file for mLLMCelltype manuscript.

Contains hierarchical metrics (hP, hR, hF1) per dataset for all methods,
macro-averaged summary, and sensitivity analysis.
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
from openpyxl.utils import get_column_letter
import re

# ============================================================
# Paths
# ============================================================
HIER_PER_DATASET = "analysis/benchmark/hierarchical_metrics_per_dataset.csv"
NAIVE_BASELINE = "results/benchmark/reference_comparison/2_evaluation/naive_baseline_comparison.csv"
SINGLER_SCTYPE = "results/benchmark/singler_sctype_batch/singler_sctype_co_matched_results.csv"
MACRO_SUMMARY = "analysis/benchmark/hierarchical_metrics_macro_summary.csv"
SENSITIVITY = "analysis/benchmark/hierarchical_metrics_sensitivity.csv"
OUTPUT = "manuscript/Supplementary_Data_2.xlsx"

# Depth ratio for hierarchical metrics (from compute_hierarchical_metrics.py)
DEPTH_RATIO = 0.70

# ============================================================
# Dataset category classification
# ============================================================
def classify_dataset(name):
    """Classify a dataset into its category for sorting."""
    n = name.strip()
    if "(Azimuth)" in n or n.lower().endswith("(azimuth)"):
        return "Azimuth"
    if "(TS)" in n or n.lower().endswith("(ts)"):
        return "TS"
    # Check for Literature + GTEx in any format (handles both
    # "Breast(Literature)(GTEx)" and "esophagus(Literature (GTEx))")
    if "Literature" in n and "GTEx" in n:
        return "GTEx Literature"
    if "(GTEx)" in n or "GTEx" in n:
        return "GTEx DE"
    if "(Cancer)" in n:
        return "Cancer"
    # HCL, MCA, etc.
    return "Other"

CATEGORY_ORDER = {
    "Azimuth": 0,
    "TS": 1,
    "GTEx DE": 2,
    "GTEx Literature": 3,
    "Cancer": 4,
    "Other": 5,
}

# ============================================================
# Styles
# ============================================================
HEADER_FONT = Font(name="Calibri", bold=True, size=11)
TITLE_FONT = Font(name="Calibri", bold=True, size=13)
NORMAL_FONT = Font(name="Calibri", size=11)
HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
THIN_BORDER = Border(
    bottom=Side(style="thin", color="B0B0B0"),
)
HEADER_BORDER = Border(
    bottom=Side(style="medium", color="000000"),
)
NUM_FMT_4 = "0.0000"


def auto_width(ws, min_width=10, max_width=30):
    """Auto-adjust column widths based on content."""
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        max_len = 0
        for cell in col_cells:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        width = min(max(max_len + 2, min_width), max_width)
        ws.column_dimensions[col_letter].width = width


def style_header_row(ws, row, ncols):
    """Style a header row with bold font, fill, and border."""
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = HEADER_BORDER
        cell.alignment = Alignment(horizontal="center", wrap_text=True)


def write_title(ws, title, ncols):
    """Write a merged title row at row 1."""
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    cell = ws.cell(row=1, column=1)
    cell.value = title
    cell.font = TITLE_FONT
    cell.alignment = Alignment(horizontal="left")


# ============================================================
# Load data
# ============================================================
print("Loading data...")

# Per-dataset hierarchical metrics (mLLMCelltype & GPTCelltype)
hier_df = pd.read_csv(HIER_PER_DATASET)

# Naive baseline
naive_df = pd.read_csv(NAIVE_BASELINE)

# Macro summary
macro_df = pd.read_csv(MACRO_SUMMARY)

# Sensitivity
sens_df = pd.read_csv(SENSITIVITY)

# ============================================================
# Sheet 1: Per-Dataset Metrics
# ============================================================
print("Building Sheet 1: Per-Dataset Metrics...")

# Pivot hierarchical metrics: one row per dataset with method columns
mllm = hier_df[hier_df["method"] == "mLLMCelltype"].copy()
gpt = hier_df[hier_df["method"] == "GPTCelltype"].copy()

mllm = mllm.rename(columns={
    "accuracy": "mLLMCelltype_Accuracy",
    "hP": "mLLMCelltype_hP",
    "hR": "mLLMCelltype_hR",
    "hF1": "mLLMCelltype_hF1",
    "n": "N_clusters",
})
gpt = gpt.rename(columns={
    "accuracy": "GPTCelltype_Accuracy",
    "hP": "GPTCelltype_hP",
    "hR": "GPTCelltype_hR",
    "hF1": "GPTCelltype_hF1",
})

mllm_cols = ["dataset", "N_clusters", "mLLMCelltype_Accuracy", "mLLMCelltype_hP", "mLLMCelltype_hR", "mLLMCelltype_hF1"]
gpt_cols = ["dataset", "GPTCelltype_Accuracy", "GPTCelltype_hP", "GPTCelltype_hR", "GPTCelltype_hF1"]

merged = mllm[mllm_cols].merge(gpt[gpt_cols], on="dataset", how="outer")

# Match naive baseline -- need to normalize dataset names
# Differences: underscores vs spaces, case, "Colon" vs "CRC",
# "esophagus(Literature (GTEx))" vs "esophagus(Literature)(GTEx)"
def normalize_dataset_name(name):
    """Normalize dataset name for matching: lowercase, remove underscores/spaces, standardize parentheses."""
    n = name.lower().strip()
    n = n.replace("_", " ")
    # Standardize "literature (gtex)" -> "literature)(gtex)" pattern
    n = n.replace("(literature (gtex))", "(literature)(gtex)")
    # Map CRC -> Colon
    n = n.replace("crc(cancer)", "colon(cancer)")
    # Remove all spaces for robust matching
    n = re.sub(r"\s+", "", n)
    return n

naive_df["dataset_key"] = naive_df["Dataset"].apply(normalize_dataset_name)
merged["dataset_key"] = merged["dataset"].apply(normalize_dataset_name)

naive_sub = naive_df[["dataset_key", "Naive_Accuracy"]].copy()
merged = merged.merge(naive_sub, on="dataset_key", how="left")

# For naive baseline: hP = hR = hF1 = accuracy
merged["Naive_hF1"] = merged["Naive_Accuracy"]

# Drop helper column
merged.drop(columns=["dataset_key"], inplace=True)

# ---- SingleR / scType (direction-aware hierarchical metrics) ----
import os, json, csv, glob as globmod

PREDICTIONS_DIR = "results/benchmark/singler_sctype_batch"
CACHE_FILES = [
    os.path.join(PREDICTIONS_DIR, "match_score_cache.json"),
    os.path.join(PREDICTIONS_DIR, "match_score_cache_local.json"),
]

# Dataset name mapping (arseven filename -> benchmark name)
PRED_DATASET_NAME_MAP = {
    "BCL": "BCL(Cancer)", "BladderTS": "Bladder(TS)", "BloodTS": "Blood(TS)",
    "Bone_MarrowTS": "Bone Marrow(TS)", "EyeTS": "Eye(TS)", "FatTS": "Fat(TS)",
    "HeartTS": "Heart(TS)", "Large_IntestineTS": "Large Intestine(TS)",
    "LiverTS": "Liver(TS)", "LungTS": "Lung(TS)", "Lymph_NodeTS": "Lymph Node(TS)",
    "MammaryTS": "Mammary(TS)", "MuscleTS": "Muscle(TS)", "PancreasTS": "Pancreas(TS)",
    "ProstateTS": "Prostate(TS)", "Salivary_GlandTS": "Salivary Gland(TS)",
    "SkinTS": "Skin(TS)", "Small_IntestineTS": "Small Intestine(TS)",
    "SpleenTS": "Spleen(TS)", "ThymusTS": "Thymus(TS)", "TongueTS": "Tongue(TS)",
    "TracheaTS": "Trachea(TS)", "UterusTS": "Uterus(TS)", "VasculatureTS": "Vasculature(TS)",
    "Mammary(TS)": "Mammary(TS)", "Pancreas(TS)": "Pancreas(TS)",
    "Lung_Cancer": "Lung(Cancer)", "HCL": "HCL", "CRC": "Colon(Cancer)",
    "InfantGut": "InfantGut",
    # GTEx snRNA-seq atlas (Eraslan et al. 2022)
    # Both esophagus tissues map to "Esophagus(GTEx)" to match benchmark naming
    "GTEx_breast": "Breast(GTEx)",
    "GTEx_esophagus_mucosa": "Esophagus(GTEx)",
    "GTEx_esophagus_muscularis": "Esophagus(GTEx)",
    "GTEx_heart": "Heart(GTEx)", "GTEx_lung": "Lung(GTEx)",
    "GTEx_prostate": "Prostate(GTEx)", "GTEx_skeletal_muscle": "Skeletal Muscle(GTEx)",
    "GTEx_skin": "Skin(GTEx)",
    # GTEx Literature variants (same expression data, same SingleR/scType predictions)
    "GTEx_breast_literature": "Breast(Literature)(GTEx)",
    "GTEx_heart_literature": "Heart(Literature)(GTEx)",
    "GTEx_lung_literature": "Lung(Literature)(GTEx)",
    "GTEx_prostate_literature": "Prostate(Literature)(GTEx)",
    "GTEx_skeletal_muscle_literature": "Skeletal Muscle(Literature)(GTEx)",
    "GTEx_skin_literature": "Skin(Literature)(GTEx)",
    "GTEx_esophagus_mucosa_literature": "esophagus(Literature (GTEx))",
    "GTEx_esophagus_muscularis_literature": "esophagus(Literature (GTEx))",
    # Azimuth datasets (CELLxGENE expression data)
    "Azimuth_adipose": "Adipose(Azimuth)",
    "Azimuth_bone_marrow": "Bone Marrow(Azimuth)",
    "Azimuth_fetal": "Fetal Development(Azimuth)",
    "Azimuth_heart": "Heart(Azimuth)",
    "Azimuth_kidney": "Kidney(Azimuth)",
    "Azimuth_liver": "Liver(Azimuth)",
    "Azimuth_lung": "Lung(Azimuth)",
    "Azimuth_pancreas": "Pancreas(Azimuth)",
    # MCA (Mouse Cell Atlas)
    "MCA": "MCA",
}

# Load match score caches
match_cache = {}
for cpath in CACHE_FILES:
    if os.path.exists(cpath):
        with open(cpath) as f:
            match_cache.update(json.load(f))

def get_cached_score(pred, ref):
    """Look up match score: cache key is pred.lower()|||ref.lower()"""
    key = f"{pred.lower()}|||{ref.lower()}"
    return match_cache.get(key, 0.0)

def determine_partial_direction(reference, prediction):
    """Heuristically determine if prediction is more general or more specific."""
    ref_lower = reference.lower().strip()
    pred_lower = prediction.lower().strip()
    stop_words = {"cell", "cells", "type", "the", "of", "and", "a", "an"}
    ref_tokens = [w for w in ref_lower.split() if w not in stop_words]
    pred_tokens = [w for w in pred_lower.split() if w not in stop_words]
    if len(pred_tokens) < len(ref_tokens):
        return "pred_more_general"
    elif len(pred_tokens) > len(ref_tokens):
        return "pred_more_specific"
    elif len(pred_lower) < len(ref_lower):
        return "pred_more_general"
    elif len(pred_lower) > len(ref_lower):
        return "pred_more_specific"
    return "symmetric"

# Compute direction-aware hierarchical metrics from prediction files
singler_metrics = {}
sctype_metrics = {}
pred_files = sorted(globmod.glob(os.path.join(PREDICTIONS_DIR, "*_predictions.csv")))
print(f"  Loaded SingleR/scType data: {len(pred_files)} prediction files")

# Accumulate rows per dataset (handles multiple files mapping to same name,
# e.g. esophagus_mucosa + esophagus_muscularis -> Esophagus(GTEx))
dataset_rows = {}
for fpath in pred_files:
    fname = os.path.basename(fpath)
    ds_key = fname.replace("_predictions.csv", "")
    ds_display = PRED_DATASET_NAME_MAP.get(ds_key, ds_key)

    with open(fpath) as f:
        reader = csv.DictReader(f)
        rows = list(reader)

    if not rows:
        continue
    if ds_display not in dataset_rows:
        dataset_rows[ds_display] = []
    dataset_rows[ds_display].extend(rows)

print(f"  Mapped to {len(dataset_rows)} unique datasets")

# Per-dataset direction-aware computation
for ds_display, rows in dataset_rows.items():
    for method_name, pred_col, metrics_dict in [
        ("SingleR", "singler_prediction", singler_metrics),
        ("scType", "sctype_prediction", sctype_metrics),
    ]:
        hp_sum, hr_sum, n, acc_sum = 0.0, 0.0, 0, 0.0
        for row in rows:
            ref = row.get("reference", "").strip()
            pred = row.get(pred_col, "").strip()
            if not ref or not pred:
                continue
            score = get_cached_score(pred, ref)
            n += 1
            acc_sum += score
            if score == 1.0:
                hp_sum += 1.0
                hr_sum += 1.0
            elif score == 0.5:
                direction = determine_partial_direction(ref, pred)
                if direction == "pred_more_general":
                    hp_sum += 1.0
                    hr_sum += DEPTH_RATIO
                elif direction == "pred_more_specific":
                    hp_sum += DEPTH_RATIO
                    hr_sum += 1.0
                else:
                    hp_sum += (1.0 + DEPTH_RATIO) / 2
                    hr_sum += (1.0 + DEPTH_RATIO) / 2

        if n > 0:
            hp = round(hp_sum / n, 4)
            hr = round(hr_sum / n, 4)
            hf1 = round(2 * hp * hr / (hp + hr), 4) if (hp + hr) > 0 else 0.0
            acc = round(acc_sum / n, 4)
            metrics_dict[ds_display] = {
                f"{method_name}_Accuracy": acc,
                f"{method_name}_hP": hp,
                f"{method_name}_hR": hr,
                f"{method_name}_hF1": hf1,
            }

# Merge SingleR/scType into main dataframe
for col in ["SingleR_Accuracy", "SingleR_hP", "SingleR_hR", "SingleR_hF1",
             "scType_Accuracy", "scType_hP", "scType_hR", "scType_hF1"]:
    merged[col] = np.nan

for idx, row in merged.iterrows():
    ds = row["dataset"]
    if ds in singler_metrics:
        for k, v in singler_metrics[ds].items():
            merged.at[idx, k] = v
    if ds in sctype_metrics:
        for k, v in sctype_metrics[ds].items():
            merged.at[idx, k] = v

n_sr_matched = merged["SingleR_Accuracy"].notna().sum()
n_sc_matched = merged["scType_Accuracy"].notna().sum()
print(f"  SingleR matched: {n_sr_matched}/{len(merged)}")
print(f"  scType matched: {n_sc_matched}/{len(merged)}")

# Classify and sort
merged["Category"] = merged["dataset"].apply(classify_dataset)
merged["cat_order"] = merged["Category"].map(CATEGORY_ORDER)
merged = merged.sort_values(
    by=["cat_order", "mLLMCelltype_Accuracy"],
    ascending=[True, False],
).reset_index(drop=True)

# Rename for output
merged = merged.rename(columns={"dataset": "Dataset"})

# Final column order for Sheet 1
sheet1_cols = [
    "Dataset", "Category", "N_clusters",
    "mLLMCelltype_Accuracy", "mLLMCelltype_hP", "mLLMCelltype_hR", "mLLMCelltype_hF1",
    "GPTCelltype_Accuracy", "GPTCelltype_hP", "GPTCelltype_hR", "GPTCelltype_hF1",
    "SingleR_Accuracy", "SingleR_hP", "SingleR_hR", "SingleR_hF1",
    "scType_Accuracy", "scType_hP", "scType_hR", "scType_hF1",
    "Naive_Accuracy", "Naive_hF1",
]
sheet1_data = merged[sheet1_cols]

# ============================================================
# Sheet 2: Macro Summary
# ============================================================
print("Building Sheet 2: Macro Summary...")

# Compute naive summary from sheet1 data
naive_acc = sheet1_data["Naive_Accuracy"].dropna()
naive_hf1 = sheet1_data["Naive_hF1"].dropna()

naive_row = {
    "Method": "Naive baseline",
    "N_datasets": int(naive_acc.count()),
    "Accuracy_mean": naive_acc.mean(),
    "Accuracy_std": naive_acc.std(),
    "hP_mean": naive_acc.mean(),  # hP = accuracy for naive
    "hP_std": naive_acc.std(),
    "hR_mean": naive_acc.mean(),
    "hR_std": naive_acc.std(),
    "hF1_mean": naive_hf1.mean(),
    "hF1_std": naive_hf1.std(),
}

# Combine with existing macro summary
sheet2_rows = []
for _, row in macro_df.iterrows():
    sheet2_rows.append({
        "Method": row["Method"],
        "N_datasets": int(row["N_datasets"]),
        "Accuracy_mean": row["Accuracy_mean"],
        "Accuracy_std": row["Accuracy_std"],
        "hP_mean": row["hP_mean"],
        "hP_std": row["hP_std"],
        "hR_mean": row["hR_mean"],
        "hR_std": row["hR_std"],
        "hF1_mean": row["hF1_mean"],
        "hF1_std": row["hF1_std"],
    })
# SingleR row (from per-dataset)
sr_acc = sheet1_data["SingleR_Accuracy"].dropna()
sr_hp = sheet1_data["SingleR_hP"].dropna()
sr_hr = sheet1_data["SingleR_hR"].dropna()
sr_hf1 = sheet1_data["SingleR_hF1"].dropna()
if len(sr_acc) > 0:
    sheet2_rows.append({
        "Method": "SingleR",
        "N_datasets": int(sr_acc.count()),
        "Accuracy_mean": sr_acc.mean(), "Accuracy_std": sr_acc.std(),
        "hP_mean": sr_hp.mean(), "hP_std": sr_hp.std(),
        "hR_mean": sr_hr.mean(), "hR_std": sr_hr.std(),
        "hF1_mean": sr_hf1.mean(), "hF1_std": sr_hf1.std(),
    })

# scType row (from per-dataset)
sc_acc = sheet1_data["scType_Accuracy"].dropna()
sc_hp = sheet1_data["scType_hP"].dropna()
sc_hr = sheet1_data["scType_hR"].dropna()
sc_hf1 = sheet1_data["scType_hF1"].dropna()
if len(sc_acc) > 0:
    sheet2_rows.append({
        "Method": "scType",
        "N_datasets": int(sc_acc.count()),
        "Accuracy_mean": sc_acc.mean(), "Accuracy_std": sc_acc.std(),
        "hP_mean": sc_hp.mean(), "hP_std": sc_hp.std(),
        "hR_mean": sc_hr.mean(), "hR_std": sc_hr.std(),
        "hF1_mean": sc_hf1.mean(), "hF1_std": sc_hf1.std(),
    })

sheet2_rows.append(naive_row)

sheet2_data = pd.DataFrame(sheet2_rows)

# Reorder: mLLMCelltype first
method_order = {"mLLMCelltype": 0, "GPTCelltype": 1, "SingleR": 2, "scType": 3, "Naive baseline": 4}
sheet2_data["_order"] = sheet2_data["Method"].map(method_order)
sheet2_data = sheet2_data.sort_values("_order").drop(columns=["_order"]).reset_index(drop=True)

# ============================================================
# Sheet 3: Sensitivity Analysis
# ============================================================
print("Building Sheet 3: Sensitivity Analysis...")
sheet3_data = sens_df.copy()

# ============================================================
# Write Excel
# ============================================================
print("Writing Excel file...")
wb = Workbook()

# --- Sheet 1 ---
ws1 = wb.active
ws1.title = "Per-Dataset Metrics"

title1 = "Supplementary Data 2 - Per-Dataset Hierarchical Metrics for All Methods"
ncols1 = len(sheet1_cols)
write_title(ws1, title1, ncols1)

# Headers at row 2
for c, col_name in enumerate(sheet1_cols, 1):
    ws1.cell(row=2, column=c, value=col_name)
style_header_row(ws1, 2, ncols1)

# Data starting at row 3
metric_cols_1 = {
    "mLLMCelltype_Accuracy", "mLLMCelltype_hP", "mLLMCelltype_hR", "mLLMCelltype_hF1",
    "GPTCelltype_Accuracy", "GPTCelltype_hP", "GPTCelltype_hR", "GPTCelltype_hF1",
    "SingleR_Accuracy", "SingleR_hP", "SingleR_hR", "SingleR_hF1",
    "scType_Accuracy", "scType_hP", "scType_hR", "scType_hF1",
    "Naive_Accuracy", "Naive_hF1",
}

for r_idx, (_, row) in enumerate(sheet1_data.iterrows(), start=3):
    for c_idx, col_name in enumerate(sheet1_cols, 1):
        cell = ws1.cell(row=r_idx, column=c_idx)
        val = row[col_name]
        if pd.isna(val):
            cell.value = None
        elif col_name == "N_clusters":
            cell.value = int(val)
        else:
            cell.value = val
        cell.font = NORMAL_FONT
        if col_name in metric_cols_1 and not pd.isna(val):
            cell.number_format = NUM_FMT_4
        cell.border = THIN_BORDER

auto_width(ws1)

# --- Sheet 2 ---
ws2 = wb.create_sheet("Macro Summary")

title2 = "Supplementary Data 2 - Macro-Averaged Summary Across 49 Datasets"
sheet2_cols = list(sheet2_data.columns)
ncols2 = len(sheet2_cols)
write_title(ws2, title2, ncols2)

for c, col_name in enumerate(sheet2_cols, 1):
    ws2.cell(row=2, column=c, value=col_name)
style_header_row(ws2, 2, ncols2)

metric_cols_2 = {
    "Accuracy_mean", "Accuracy_std", "hP_mean", "hP_std",
    "hR_mean", "hR_std", "hF1_mean", "hF1_std",
}

for r_idx, (_, row) in enumerate(sheet2_data.iterrows(), start=3):
    for c_idx, col_name in enumerate(sheet2_cols, 1):
        cell = ws2.cell(row=r_idx, column=c_idx)
        val = row[col_name]
        if pd.isna(val):
            cell.value = None
        elif col_name == "N_datasets":
            cell.value = int(val)
        else:
            cell.value = val
        cell.font = NORMAL_FONT
        if col_name in metric_cols_2 and not pd.isna(val):
            cell.number_format = NUM_FMT_4
        cell.border = THIN_BORDER

auto_width(ws2)

# --- Sheet 3 ---
ws3 = wb.create_sheet("Sensitivity Analysis")

title3 = "Supplementary Data 2 - Hierarchical Metrics Sensitivity to depth_ratio Parameter"
sheet3_cols = list(sheet3_data.columns)
ncols3 = len(sheet3_cols)
write_title(ws3, title3, ncols3)

for c, col_name in enumerate(sheet3_cols, 1):
    ws3.cell(row=2, column=c, value=col_name)
style_header_row(ws3, 2, ncols3)

metric_cols_3 = {"hP", "hR", "hF1", "accuracy", "depth_ratio"}

for r_idx, (_, row) in enumerate(sheet3_data.iterrows(), start=3):
    for c_idx, col_name in enumerate(sheet3_cols, 1):
        cell = ws3.cell(row=r_idx, column=c_idx)
        val = row[col_name]
        if pd.isna(val):
            cell.value = None
        elif col_name == "n":
            cell.value = int(val)
        else:
            cell.value = val
        cell.font = NORMAL_FONT
        if col_name in metric_cols_3 and not pd.isna(val):
            cell.number_format = NUM_FMT_4
        cell.border = THIN_BORDER

auto_width(ws3)

# ============================================================
# Save
# ============================================================
wb.save(OUTPUT)
print(f"Saved: {OUTPUT}")
print(f"  Sheet 1: {len(sheet1_data)} datasets")
print(f"  Sheet 2: {len(sheet2_data)} methods")
print(f"  Sheet 3: {len(sheet3_data)} rows")
